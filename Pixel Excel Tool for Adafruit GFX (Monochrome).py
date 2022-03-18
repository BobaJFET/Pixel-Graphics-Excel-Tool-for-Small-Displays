#a helpful article I used as reference for openpyxl: https://www.geeksforgeeks.org/reading-excel-file-using-python/
#Adafruit GFX website: https://learn.adafruit.com/adafruit-gfx-graphics-library/overview

import openpyxl

#Insert exact excel file location in the quotes
path = "D:/OneDrive - NJIT/ECE 414, Senior Design/Code/Adafruit GFX Excel Tool/128x64.xlsx" 

#Opening a workbook and creating a workbook object
wb_obj = openpyxl.load_workbook(path, data_only = True)

sheet_obj = wb_obj.active

#Get max rows and columns
row = sheet_obj.max_row
column = sheet_obj.max_column

displayObject = "display" #Name this according to your display hardware's driver library. For example, Adafruit_SSD1325.h uses 'display'



#Uncomment to list each pixel's on/off state and their coordinate

# for y in range(1, row + 1):
   
#     print("\nPixels in row ",y - 1 )
#     for x in range(1, column + 1):
        
#         cell_obj = sheet_obj.cell(row = y, column = x)
#         color = cell_obj.fill.fgColor
        
#         color_in_hex = cell_obj.fill.start_color.index #color in ARGB hex
#         print ('HEX =',color_in_hex) #prints the color in ARGB hex
        
#         if color.rgb != '00000000' : 
#             print("On: (", y -1 ,",", x - 1,")") 
            
#         else:
#             print("Off: (", y - 1,",", x - 1, ")")
            
#prints the number of rows and columns in the excel file. Useful to know the dimensions of your graphic. You could also draw a border around your display 
#dimensions.. 128x64 for example. 
print("\n")
print("Total Rows:", row)
print("Total Columns:", column)
print("\n")            
for y in range(1, row + 1):
    for x in range(1, column + 1):
        cell_obj = sheet_obj.cell(row = y, column = x)
        color = cell_obj.fill.bgColor
        if color.rgb != '00000000' : 
            print(displayObject,".","drawPixel(",x - 1,",",y - 1,",",1,")",";") 
            
     

