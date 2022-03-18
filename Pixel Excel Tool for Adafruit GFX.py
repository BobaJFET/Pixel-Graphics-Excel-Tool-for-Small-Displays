#a helpful article I used as reference for openpyxl: https://www.geeksforgeeks.org/reading-excel-file-using-python/
#Adafruit GFX website: https://learn.adafruit.com/adafruit-gfx-graphics-library/overview

import openpyxl

#Insert exact excel file location in the quotes
path = "" 

#Opening a workbook and creating a workbook object
wb_obj = openpyxl.load_workbook(path, data_only = True)

sheet_obj = wb_obj.active

#Get max rows and columns
row = sheet_obj.max_row
column = sheet_obj.max_column

displayObject = "" #Name this according to your display hardware's driver library. For example, Adafruit_SSD1325.h uses 'display'



#Uncomment to list each pixel's on/off state and their coordinate

# for y in range(1, row + 1):
   
#     print("\nPixels in row ",y - 1 )
#     for x in range(1, column + 1):
        
#         cell_obj = sheet_obj.cell(row = y, column = x)
#         color = cell_obj.fill.fgColor
        
#         color_in_hex = cell_obj.fill.start_color.index #color in ARGB hex
#         print ('HEX =',color_in_hex) #prints the color in ARGB hex
        
#         if color.rgb != '00000000' : 
#             print("On: (", y -1 ,",", x - 1,")") 
            
#         else:
#             print("Off: (", y - 1,",", x - 1, ")")
            
#print number of rows and columns in the excel file 
print("\n")
print("Total Rows:", row)
print("Total Columns:", column)
print("\n")            
newLine = True

#Looping through each row and column. Get the cell color for each successive cell. 
for y in range(1, row + 1):
    for x in range(1, column + 1):
        cell_obj = sheet_obj.cell(row = y, column = x)
        color = cell_obj.fill.bgColor
        
        nextCell = sheet_obj.cell(row = y, column = x+1)
        nextCellColor = nextCell.fill.bgColor
        
       
       
        if color.rgb != '00000000' and newLine == True:  #Check if a pixel (or cell) is 'on' (has color other than nofill which is '00000000')
            x0 = x                                       #also checks if we are on a new line of adjacent horiztonal pixels
            newLine = False
            if nextCellColor.rgb == '00000000': #If the next pixel is off, then this is not a line of pixels, and we'll just draw the individual pixel
                print(displayObject,".drawPixel(",x - 1,",",y - 1,",1);")
                newLine = True
            
            

        elif color.rgb != '00000000' and newLine == False: #if our next pixel is on, and adjacent to the last one, we'll update our x1 coordinate
                x1 = x
               
        elif color.rgb == '00000000' and newLine == False: #once we encounter an 'off' pixel, our line has ended, and we draw the line from x0 to x1
            print(displayObject,".drawLine(",x0 - 1,",", y - 1, ",",x1 - 1,",",y - 1,",1)")
            newLine = True
                        
